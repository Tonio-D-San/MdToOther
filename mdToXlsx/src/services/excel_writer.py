import openpyxl

from .formatter import (
    style_header,
    style_cell,
    apply_status_color,
    resize_columns,
)


def write_tables_to_excel(excel_input, excel_output, sheet_name, start_row, tables):
    wb = openpyxl.load_workbook(excel_input)
    ws = wb[sheet_name]

    for df in tables:
        # Header
        for j, header in enumerate(df.columns):
            style_header(ws.cell(row=start_row, column=1 + j, value=header))

        # Body
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                cell = ws.cell(row=start_row + 1 + i, column=1 + j, value=value)
                style_cell(cell)
                apply_status_color(df.columns[j], value, cell)

        start_row += len(df) + 3

    resize_columns(ws)
    wb.save(excel_output)
