from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..domain.table_colors import STATUS_COLORS

_header_font = Font(bold=True, color="FFFFFF")
_header_fill = PatternFill("solid", fgColor="4F81BD")
_center_align = Alignment(horizontal="center", vertical="center")
_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(cell):
    cell.font = _header_font
    cell.fill = _header_fill
    cell.alignment = _center_align
    cell.border = _border

def style_cell(cell):
    cell.alignment = _center_align
    cell.border = _border

def apply_status_color(column, value, cell):
    if column.lower() != "migrato":
        return

    if value in STATUS_COLORS:
        cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[value])

def resize_columns(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 5
